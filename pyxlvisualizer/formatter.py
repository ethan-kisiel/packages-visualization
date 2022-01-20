from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .color import Color

class Formatter:
    def __init__(self, keys=None, colors=None):
        if colors is None:
            colors = [Color([255, 255, 255, 0.5]).get_argb_hex(),
                      Color([0, 255, 255, 0.5]).get_argb_hex(),
                      Color([255, 0, 255, 0.5]).get_argb_hex(),
                      Color([255, 255, 0, 0.5]).get_argb_hex()]
        if keys is None:    
            keys = ['A', 'B', 'C', 'D']
            
        # TODO: add exception for mismatched array size
        self.colors = dict(zip(keys, colors))

    # Overwrites the color-code dictionary
    def overwrite_colors(self, new_dict: dict):
        self.colors = new_dict
    
    # Color must be ARGB hex value without '#'
    def set_color(self, key, color: str):
        try:
            self.colors[key] = color
        except KeyError:
            print('KeyError; invalid key')
    
    def set_key(self, key, new_key):
        try:
            self.colors[new_key] = self.colors.pop(key)
        except KeyError:
            print('KeyError; invalid key')

    # TODO: Adjust file_loc/file_name to make renaming of file smoother
    # TODO: Add functionality for changing fill type
    def format_spreadsheet(self, file_loc: str, new_sheet_name: str=None, cell_columns=None):
        '''
        Takes file location and optional variables for new sheet name and cell columns
        Uses new sheet and cell columns to format coloring with self.colors based off self.keys
        '''
        if cell_columns is None:
            cell_columns = ['A', 'B', 'C']
        if new_sheet_name is None:
            new_sheet_name = 'New Sheet'

        wb = load_workbook(filename=file_loc)
        ws = wb.active
        counter = 2
        # while first column, row HAS data
        while ws[f'{cell_columns[0]}{counter}'].internal_value:
            # get the value of first column @ current row (counter)
            current_person = ws[f'A{counter}'].internal_value
            #print(color_codes[current_person])
            for i,_ in enumerate(cell_columns):
                ws[f'{cell_columns[i]}{counter}'].fill = PatternFill(fill_type='lightUp', fgColor=self.colors[current_person])
            counter += 1
        ws.title = new_sheet_name
        wb.save(filename=f'NEW-{file_loc}')

    def __repr__(self):
        these_keys = ''
        for key in self.colors:
            these_keys += f'{key}:{self.colors[key]},'
            
        return these_keys
        